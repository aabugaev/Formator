#!/usr/bin/env python
# coding: utf-8
import subprocess
import sys

def install(package):
    subprocess.call([sys.executable, "-m", "pip", "install", package])

try:
    import pandas as pd
    import numpy as np
    import xlrd

except:
    install('pandas')
    install('numpy')
    install('xlrd')



# In[11]:


import pandas as pd
import numpy as np
import itertools
from math import isnan
from itertools import zip_longest
import os
import re

if not os.path.exists("Forms"):
    os.mkdir("Forms")


# In[12]:


Groups_File_Name = str(input("Please give the name of the file with Items.\n"))


# In[13]:


Groups_df = pd.read_excel(Groups_File_Name, dtype=str)


# In[14]:


Groups_dict = Groups_df.to_dict("list")

#cleaning the list from nans
for Group in Groups_dict:
    Groups_dict[Group] = [str(word) for word in Groups_dict[Group] if str(word) != 'nan']
    print(Groups_dict[Group])


# In[15]:


Groups_dict


# In[16]:


Forms_File_Name = str(input("Please give the name of the file with forms.\n"))


# In[17]:


Forms_df = pd.read_excel(Forms_File_Name, dtype=str)
Forms_List = Forms_df["Forms"].tolist()


# In[18]:


for form in Forms_List:
    Groups_dict= Groups_df[form.split(" ")].to_dict("list")
    
    for Group in Groups_dict:
        Groups_dict[Group] = [str(word) for word in Groups_dict[Group] if str(word) != 'nan']
        
    combinations = itertools.product(*Groups_dict.values())
    combinations_joined = list(map(lambda x: " ".join(x), list(combinations)))
    print(combinations_joined)
    
    combinations_df = pd.DataFrame(combinations_joined) 
    combinations_df.to_excel("Forms\\" + form + ".xlsx")


# In[ ]:





# In[19]:


#from datetime import datetime
#timestamp= datetime.now().strftime("%Y%m%d-%H%M%S")
#Label_df.to_excel("Labelling\\Label_"+timestamp+"_"+Label_File_name)


# In[20]:


"""

Без кавычек:

#%load file.py
%%writefile file.py  - в начале блока
%pycat  -
%run file.py
%lsmagic

from IPython.core.interactiveshell import InteractiveShell
InteractiveShell.ast_node_interactivity = "all"

===openpyxl===
from openpyxl import load_workbook
from openpyxl import Workbook

wb = load_workbook()
wb_ws = wb.get_active_sheet()

wrwb = Workbook()
wrwb_ws = wrwb.get_active_sheet()

wb.save()

===numpy/pandas===
import pandas as pd
import numpy as np

excel_df = pd.read_excel()
csv_df = pd.read_csv()


df.to_excel()
df.to_csv()

writer = pd.ExcelWriter('',engine='xlsxwriter',options={})
df.to_excel(writer)
writer.save()


====requests/BeautifulSoup===
import requests
from bs4 import BeautifulSoup

page = requests.get("http://yandex.ru")
page.encoding = "windows-1251"
soup = BeautifulSoup(''.join(page.text), "html.parser\"),
soup.findAll("div")


===Files and directories===
import os
FileList = os.listdir()

#if not os.path.exists("Folder"):
#   os.mkdir("Folder") 

"""

